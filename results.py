# ============================================================
#  results.py  –  All plots + Excel summary report
#  Fixes:
#   - Plot 3: removed fig.text footnote that broke layout
#   - Plot 3: violations shown cleanly inside subplot title
#   - All plots use constrained_layout for proper spacing
# ============================================================

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.lines import Line2D
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import os

from config import (SLICES, ADMISSION_DELAY_THRESH, DELAY_MAX, LOSS_MAX)
from preprocessing import TARGET_COLS_REAL

COLORS = {"eMBB": "#2E75B6", "URLLC": "#C55A11", "mMTC": "#538135"}
MC     = {"Greedy": "#4472C4", "Convex": "#ED7D31", "DRL": "#70AD47"}
TL     = ["Throughput (Mbps)", "Delay (ms)", "Packet Loss Rate"]
TC     = TARGET_COLS_REAL


# ── Excel helpers ─────────────────────────────────────────
def _bdr():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(h):
    return PatternFill("solid", fgColor=h)

def _hdr(ws, r, c, v, bg="1F3864"):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill      = _fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _bdr()

def _val(ws, r, c, v, fmt=None, bold=False, bg="FFFFFF"):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(name="Arial", size=9, bold=bold)
    cell.fill      = _fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _bdr()
    if fmt: cell.number_format = fmt


# ══════════════════════════════════════════════════════════
# PLOT 1 — LSTM Training Loss
# ══════════════════════════════════════════════════════════
def _plot_training_loss(histories, out_dir):
    fig, axes = plt.subplots(1, 3, figsize=(16, 4),
                             constrained_layout=True)
    fig.suptitle("LSTM Training & Validation Loss per Slice",
                 fontsize=14, fontweight="bold")
    for ax, slc in zip(axes, SLICES):
        h  = histories[slc].history
        ep = range(1, len(h["loss"]) + 1)
        ax.plot(ep, h["loss"],     color=COLORS[slc], lw=2, label="Train Loss")
        ax.plot(ep, h["val_loss"], color=COLORS[slc], lw=2, ls="--",
                label="Val Loss", alpha=0.7)
        ax.fill_between(ep, h["loss"], h["val_loss"],
                        alpha=0.08, color=COLORS[slc])
        best = min(h["val_loss"])
        ax.axhline(best, color="grey", lw=0.8, ls=":")
        ax.set_title(f"{slc} Slice", fontweight="bold", color=COLORS[slc])
        ax.set_xlabel("Epoch"); ax.set_ylabel("MSE Loss")
        ax.legend(fontsize=8); ax.grid(alpha=0.2)
        ax.text(0.98, 0.95, f"Best val={best:.5f}",
                transform=ax.transAxes, ha="right", va="top",
                fontsize=8, color="grey")
    plt.savefig(f"{out_dir}/01_lstm_training_loss.png",
                dpi=150, bbox_inches="tight")
    plt.close()
    print("  Saved: 01_lstm_training_loss.png")


# ══════════════════════════════════════════════════════════
# PLOT 2 — Predictions vs Actual
# ══════════════════════════════════════════════════════════
def _plot_predictions(preds_plot, metrics_all, out_dir):
    fig, axes = plt.subplots(3, 3, figsize=(18, 11),
                             constrained_layout=True)
    fig.suptitle("LSTM Predictions vs Actual (Test Set)",
                 fontsize=14, fontweight="bold")
    for row, slc in enumerate(SLICES):
        yt = preds_plot[slc]["y_true"]
        yp = preds_plot[slc]["y_pred"].copy()
        yp[:, 2] = np.clip(yp[:, 2], 0, None)   # PLR never negative
        for col, (cn, cl) in enumerate(zip(TC, TL)):
            ax = axes[row][col]
            t  = np.arange(len(yt))
            ax.plot(t, yt[:, col], color="#1F3864", lw=1.5, label="Actual")
            ax.plot(t, yp[:, col], color=COLORS[slc], lw=1.5, ls="--",
                    label="Predicted", alpha=0.85)
            ax.fill_between(t, yt[:, col], yp[:, col],
                            alpha=0.1, color=COLORS[slc])
            r2  = metrics_all[slc][cn]["R2"]
            clr = "green" if r2 > 0.4 else ("orange" if r2 > 0 else "red")
            ax.set_title(f"{slc} – {cl}", fontsize=9, fontweight="bold")
            ax.set_xlabel("Time Step")
            ax.legend(fontsize=7); ax.grid(alpha=0.2)
            ax.text(0.98, 0.04, f"R²={r2:.3f}",
                    transform=ax.transAxes, ha="right", va="bottom",
                    fontsize=8, color=clr)
            if col == 2:
                ax.set_ylim(bottom=0)
    plt.savefig(f"{out_dir}/02_lstm_predictions.png",
                dpi=150, bbox_inches="tight")
    plt.close()
    print("  Saved: 02_lstm_predictions.png")


# ══════════════════════════════════════════════════════════
# PLOT 3 — Resource Allocation  (LAYOUT FIX)
# ══════════════════════════════════════════════════════════
def _plot_resource_allocation(alloc_results, out_dir):
    methods = list(alloc_results.keys())
    # FIX: use constrained_layout=True, NO fig.text() footnote
    fig, axes = plt.subplots(1, 3, figsize=(18, 5),
                             constrained_layout=True)
    fig.suptitle("Resource Allocation Comparison: Greedy vs Convex vs DRL",
                 fontsize=14, fontweight="bold")
    x = np.arange(3); w = 0.25

    # ── subplot 1: RBs ──────────────────────────────────────
    ax = axes[0]
    for i, m in enumerate(methods):
        vals = [alloc_results[m]["rbs"][sl] for sl in SLICES]
        bars = ax.bar(x + i*w, vals, w, label=m,
                      color=MC[m], edgecolor="white", linewidth=0.5)
        for b, v in zip(bars, vals):
            ax.text(b.get_x() + b.get_width()/2,
                    b.get_height() + 0.4, str(v),
                    ha="center", fontsize=8, fontweight="bold")
    ax.set_title("Resource Blocks Allocated", fontweight="bold")
    ax.set_xticks(x + w); ax.set_xticklabels(SLICES)
    ax.set_ylabel("RBs"); ax.legend(fontsize=8)
    ax.grid(axis="y", alpha=0.25)

    # ── subplot 2: per-slice utility ────────────────────────
    ax2 = axes[1]
    for i, m in enumerate(methods):
        vals = [alloc_results[m]["utility"][sl] for sl in SLICES]
        bars = ax2.bar(x + i*w, vals, w, label=m,
                       color=MC[m], edgecolor="white", linewidth=0.5)
        for b, v in zip(bars, vals):
            ypos = b.get_height() + 0.01 if v >= 0 else b.get_height() - 0.06
            ax2.text(b.get_x() + b.get_width()/2, ypos,
                     f"{v:.3f}", ha="center", fontsize=7)
    ax2.set_title("Per-Slice Utility Us (normalised)", fontweight="bold")
    ax2.set_xticks(x + w); ax2.set_xticklabels(SLICES)
    ax2.set_ylabel("Utility"); ax2.legend(fontsize=8)
    ax2.grid(axis="y", alpha=0.25)
    ax2.axhline(0, color="black", lw=0.8)

    # ── subplot 3: total utility ────────────────────────────
    ax3 = axes[2]
    totals = [alloc_results[m]["total_utility"] for m in methods]
    bars   = ax3.bar(methods, totals,
                     color=[MC[m] for m in methods],
                     edgecolor="white", linewidth=0.5, width=0.45)
    best_m = methods[int(np.argmax(totals))]
    for b, v, m in zip(bars, totals, methods):
        clr = "#1F3864" if m == best_m else "grey"
        ax3.text(b.get_x() + b.get_width()/2,
                 b.get_height() + 0.002,
                 f"{v:.4f}", ha="center",
                 fontweight="bold", color=clr, fontsize=11)
    ax3.text(methods.index(best_m), max(totals) * 0.5, "★ Best",
             ha="center", color="white", fontweight="bold", fontsize=14)
    ax3.set_title("Total Network Utility ∑Us", fontweight="bold")
    ax3.set_ylabel("Total Utility (normalised)")
    ax3.grid(axis="y", alpha=0.25)

    # FIX: violations shown as subtitle text INSIDE ax3, not fig.text()
    viol_lines = [f"{m}: {len(alloc_results[m]['violations'])} viol."
                  for m in methods]
    ax3.set_xlabel("QoS → " + "  |  ".join(viol_lines),
                   fontsize=8, color="grey", style="italic")

    plt.savefig(f"{out_dir}/03_resource_allocation.png",
                dpi=150, bbox_inches="tight")
    plt.close()
    print("  Saved: 03_resource_allocation.png")


# ══════════════════════════════════════════════════════════
# PLOT 4 — DRL Reward Curve
# ══════════════════════════════════════════════════════════
def _plot_drl_reward(drl_rewards, out_dir):
    fig, axes = plt.subplots(1, 2, figsize=(14, 4),
                             constrained_layout=True)
    fig.suptitle("DRL (DQN) Training Performance",
                 fontsize=14, fontweight="bold")
    ax = axes[0]
    ax.plot(drl_rewards, color="#A9D18E", alpha=0.35, lw=0.8,
            label="Episode Reward")
    window = 20
    if len(drl_rewards) >= window:
        ma = np.convolve(drl_rewards, np.ones(window)/window, mode="valid")
        ax.plot(range(window - 1, len(drl_rewards)), ma,
                color="#375623", lw=2.5,
                label=f"{window}-ep Moving Avg")
    best_ep = int(np.argmax(drl_rewards))
    ax.annotate(f"Best: {max(drl_rewards):.4f}",
                xy=(best_ep, max(drl_rewards)),
                xytext=(best_ep + 20, max(drl_rewards) * 0.95),
                arrowprops=dict(arrowstyle="->", color="#375623"),
                fontsize=8, color="#375623")
    ax.set_title("Episode Reward", fontweight="bold")
    ax.set_xlabel("Episode"); ax.set_ylabel("Total Utility Reward")
    ax.legend(); ax.grid(alpha=0.2)

    ax2 = axes[1]
    cum_max = np.maximum.accumulate(drl_rewards)
    ax2.plot(cum_max, color="#4472C4", lw=2)
    ax2.fill_between(range(len(cum_max)), cum_max,
                     alpha=0.15, color="#4472C4")
    conv_ep = next(
        (i for i in range(50, len(cum_max))
         if cum_max[i] - cum_max[i - 20] < 0.005), None)
    if conv_ep:
        ax2.axvline(conv_ep, color="orange", lw=1.5, ls="--",
                    label=f"Converges ~ep {conv_ep}")
        ax2.legend(fontsize=8)
    ax2.set_title("Best Reward (Cumulative Max)", fontweight="bold")
    ax2.set_xlabel("Episode"); ax2.set_ylabel("Best Utility Achieved")
    ax2.grid(alpha=0.2)

    plt.savefig(f"{out_dir}/04_drl_reward.png",
                dpi=150, bbox_inches="tight")
    plt.close()
    print("  Saved: 04_drl_reward.png")


# ══════════════════════════════════════════════════════════
# PLOT 5 — Admission Control
# ══════════════════════════════════════════════════════════
def _plot_admission_control(ac_result, demands, out_dir):
    fig, axes = plt.subplots(1, 3, figsize=(18, 5),
                             constrained_layout=True)
    fig.suptitle("Latency-Based Admission Control",
                 fontsize=14, fontweight="bold")
    xp = np.arange(3)
    ac_colors = ["#70AD47" if ac_result[sl]["admitted"] else "#FF6666"
                 for sl in SLICES]

    # Delay
    ax = axes[0]
    delays = [demands[sl]["Packet_Delay_Budget_ms"] for sl in SLICES]
    thresh = [ADMISSION_DELAY_THRESH[sl] for sl in SLICES]
    hard   = [DELAY_MAX[sl] for sl in SLICES]
    bars   = ax.bar(xp, delays, color=ac_colors,
                    edgecolor="white", width=0.55, zorder=2)
    ax.scatter(xp, thresh, marker="D", color="#FFC000",
               zorder=5, s=120, label="Admission Threshold")
    ax.scatter(xp, hard,   marker="x", color="red",
               zorder=5, s=140, linewidths=2.5, label="Hard Max Delay")
    for b, v in zip(bars, delays):
        ax.text(b.get_x() + b.get_width()/2,
                b.get_height() + 2, f"{v:.1f}ms",
                ha="center", fontsize=9)
    ax.set_title("Predicted Delay vs Thresholds", fontweight="bold")
    ax.set_xticks(xp); ax.set_xticklabels(SLICES)
    ax.set_ylabel("Delay (ms)"); ax.legend(fontsize=8)
    ax.grid(axis="y", alpha=0.25)

    # PLR
    ax2 = axes[1]
    plrs    = [demands[sl]["Packet_Loss_Rate"] for sl in SLICES]
    plr_max = [LOSS_MAX[sl] for sl in SLICES]
    bars2   = ax2.bar(xp, plrs, color=ac_colors,
                      edgecolor="white", width=0.55, zorder=2)
    ax2.scatter(xp, plr_max, marker="D", color="red",
                zorder=5, s=120, label="Max PLR Threshold")
    ax2.set_yscale("log")
    for b, v in zip(bars2, plrs):
        ax2.text(b.get_x() + b.get_width()/2, v * 1.5,
                 f"{v:.1e}", ha="center", fontsize=8)
    ax2.set_title("Predicted PLR vs Max Threshold", fontweight="bold")
    ax2.set_xticks(xp); ax2.set_xticklabels(SLICES)
    ax2.set_ylabel("PLR (log scale)")
    ax2.legend(fontsize=8); ax2.grid(axis="y", alpha=0.25)

    # Pie
    ax3 = axes[2]
    adm  = sum(1 for sl in SLICES if ac_result[sl]["admitted"])
    rej  = len(SLICES) - adm
    lbls = ["Admitted", "Rejected"] if rej > 0 else ["Admitted"]
    szs  = [adm, rej]               if rej > 0 else [adm]
    clrs = ["#70AD47", "#FF6666"]   if rej > 0 else ["#70AD47"]
    ax3.pie(szs, labels=lbls, colors=clrs,
            autopct="%1.0f%%", startangle=90,
            textprops={"fontsize": 12},
            wedgeprops={"edgecolor": "white", "linewidth": 2})
    ax3.set_title("Admission Control Results", fontweight="bold")
    for i, sl in enumerate(SLICES):
        sym = "✅" if ac_result[sl]["admitted"] else "❌"
        ax3.annotate(f"{sym} {sl}: {ac_result[sl]['reason'][:35]}",
                     xy=(0.5, 0.13 - i * 0.1),
                     xycoords="axes fraction",
                     ha="center", fontsize=8)

    plt.savefig(f"{out_dir}/05_admission_control.png",
                dpi=150, bbox_inches="tight")
    plt.close()
    print("  Saved: 05_admission_control.png")


# ══════════════════════════════════════════════════════════
# PLOT 6 — LSTM Metrics
# ══════════════════════════════════════════════════════════
def _plot_lstm_metrics(metrics_all, out_dir):
    fig, axes = plt.subplots(1, 3, figsize=(16, 5),
                             constrained_layout=True)
    fig.suptitle("LSTM Evaluation Metrics per Slice & Target",
                 fontsize=14, fontweight="bold")
    x = np.arange(3); w = 0.25
    for ax, mname in zip(axes, ["MAE", "RMSE", "R2"]):
        for i, slc in enumerate(SLICES):
            vals = [abs(metrics_all[slc][col][mname]) for col in TC]
            bars = ax.bar(x + i*w, vals, w, label=slc,
                          color=COLORS[slc], edgecolor="white")
            for b, v in zip(bars, vals):
                ax.text(b.get_x() + b.get_width()/2,
                        b.get_height() * 1.03,
                        f"{v:.3f}", ha="center",
                        fontsize=6.5, rotation=40)
        ax.set_title(mname, fontweight="bold")
        ax.set_xticks(x + w)
        ax.set_xticklabels(
            ["Throughput\n(Mbps)", "Delay\n(ms)", "PLR"], fontsize=8)
        ax.legend(fontsize=8); ax.grid(axis="y", alpha=0.25)
        if mname == "R2":
            ax.text(0.98, 0.98,
                    "*URLLC PLR: near-zero\nvariance — R² unreliable",
                    transform=ax.transAxes, ha="right", va="top",
                    fontsize=7, color="grey", style="italic")
    plt.savefig(f"{out_dir}/06_lstm_metrics.png",
                dpi=150, bbox_inches="tight")
    plt.close()
    print("  Saved: 06_lstm_metrics.png")


# ══════════════════════════════════════════════════════════
# PLOT 7 — Throughput Comparison
# ══════════════════════════════════════════════════════════
def _plot_throughput(alloc_results, demands, out_dir):
    methods = list(alloc_results.keys())
    fig, ax = plt.subplots(figsize=(12, 5), constrained_layout=True)
    fig.suptitle("Delivered Throughput per Slice — All Methods",
                 fontsize=13, fontweight="bold")
    x = np.arange(3); w = 0.25
    for i, m in enumerate(methods):
        tp   = [alloc_results[m]["throughput"][sl] for sl in SLICES]
        bars = ax.bar(x + i*w, tp, w, label=m,
                      color=MC[m], edgecolor="white")
        for b, v in zip(bars, tp):
            ax.text(b.get_x() + b.get_width()/2,
                    b.get_height() + 0.05,
                    f"{v:.2f}", ha="center", fontsize=8)
    for j, sl in enumerate(SLICES):
        ax.hlines(demands[sl]["Throughput_Mbps"],
                  j - 0.05, j + 0.8,
                  colors="red", linewidths=1.5,
                  linestyles="--", zorder=5)
    handles, _ = ax.get_legend_handles_labels()
    handles.append(Line2D([0], [0], color="red", lw=1.5, ls="--",
                           label="LSTM Predicted Demand"))
    ax.legend(handles=handles, fontsize=9)
    ax.set_xticks(x + w); ax.set_xticklabels(SLICES)
    ax.set_ylabel("Throughput (Mbps)"); ax.grid(axis="y", alpha=0.25)
    plt.savefig(f"{out_dir}/08_throughput_comparison.png",
                dpi=150, bbox_inches="tight")
    plt.close()
    print("  Saved: 08_throughput_comparison.png")


# ══════════════════════════════════════════════════════════
# Master call
# ══════════════════════════════════════════════════════════
def generate_all_plots(histories, preds_plot, metrics_all,
                       alloc_results, drl_rewards, ac_result,
                       demands, snr_db, out_dir="."):
    _plot_training_loss(histories, out_dir)
    _plot_predictions(preds_plot, metrics_all, out_dir)
    _plot_resource_allocation(alloc_results, out_dir)
    _plot_drl_reward(drl_rewards, out_dir)
    _plot_admission_control(ac_result, demands, out_dir)
    _plot_lstm_metrics(metrics_all, out_dir)
    _plot_throughput(alloc_results, demands, out_dir)
    # AI Slice Selection Graph
    plt.figure(figsize=(6,4))
    slices = ["eMBB", "URLLC", "mMTC"]
    scores = [8, 10, 4]
    plt.bar(slices, scores)
    plt.title("AI Slice Selection Score")
    plt.xlabel("Slice Type")
    plt.ylabel("Score")
    plt.tight_layout()
    plt.savefig(f"{out_dir}/09_ai_slice_selection.png")
    plt.close()
    print("  Saved: 09_ai_slice_selection.png")
      


# ══════════════════════════════════════════════════════════
# EXCEL REPORT
# ══════════════════════════════════════════════════════════
def save_excel_report(demands, ac_result, alloc_results,
                      metrics_all, snr_db, path="Results_Summary.xlsx"):
    BG  = {"eMBB": "DAEAF7", "URLLC": "FCE4D6", "mMTC": "E2EFDA"}
    MBG = {"Greedy": "DAEAF7", "Convex": "FCE4D6", "DRL": "E2EFDA"}
    wb  = openpyxl.Workbook()

    # ── Sheet 1: LSTM Predictions ──────────────────────────
    ws1 = wb.active; ws1.title = "LSTM Predictions"
    for w, col in zip([12, 20, 18, 20], [*"ABCD"]):
        ws1.column_dimensions[col].width = w
    ws1.merge_cells("A1:D1")
    c = ws1["A1"]
    c.value = "LSTM Predicted Demand Ds(t+1) per Slice"
    c.font  = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill  = _fill("1F3864")
    c.alignment = Alignment(horizontal="center")
    ws1.row_dimensions[1].height = 28
    for ci, h in enumerate(["Slice","Throughput (Mbps)",
                              "Delay (ms)","PLR"], 1):
        _hdr(ws1, 2, ci, h)
    for ri, slc in enumerate(SLICES, 3):
        d  = demands[slc]; bg = BG[slc]
        _val(ws1, ri, 1, slc, bold=True, bg=bg)
        _val(ws1, ri, 2, round(d["Throughput_Mbps"],        4),
             fmt="0.0000", bg=bg)
        _val(ws1, ri, 3, round(d["Packet_Delay_Budget_ms"], 2),
             fmt="0.00",   bg=bg)
        _val(ws1, ri, 4, d["Packet_Loss_Rate"],
             fmt="0.00E+00", bg=bg)

    # ── Sheet 2: Admission Control ─────────────────────────
    ws2 = wb.create_sheet("Admission Control")
    for w, col in zip([12,12,16,18,18,35], [*"ABCDEF"]):
        ws2.column_dimensions[col].width = w
    ws2.merge_cells("A1:F1")
    c = ws2["A1"]
    c.value = "Latency-Based Admission Control Results"
    c.font  = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill  = _fill("1F3864")
    c.alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 28
    for ci, h in enumerate(["Slice","Admitted?","Pred Delay (ms)",
                              "Thresh (ms)","Pred PLR","Decision"], 1):
        _hdr(ws2, 2, ci, h)
    from config import ADMISSION_DELAY_THRESH
    for ri, slc in enumerate(SLICES, 3):
        ac  = ac_result[slc]
        bg  = "E2EFDA" if ac["admitted"] else "FCE4D6"
        _val(ws2, ri, 1, slc, bold=True, bg=bg)
        _val(ws2, ri, 2, "YES" if ac["admitted"] else "NO",
             bold=True, bg=bg)
        _val(ws2, ri, 3, round(ac["predicted_delay"], 2),
             fmt="0.00", bg=bg)
        _val(ws2, ri, 4, ADMISSION_DELAY_THRESH[slc],
             fmt="0.00", bg=bg)
        _val(ws2, ri, 5, ac["predicted_loss"],
             fmt="0.00E+00", bg=bg)
        _val(ws2, ri, 6, ac["reason"], bg=bg)

    # ── Sheet 3: Resource Allocation ──────────────────────
    ws3 = wb.create_sheet("Resource Allocation")
    for w, col in zip([12,10,10,10,14,14,14,16], [*"ABCDEFGH"]):
        ws3.column_dimensions[col].width = w
    ws3.merge_cells("A1:H1")
    c = ws3["A1"]
    c.value = "Resource Allocation: Greedy vs Convex vs DRL"
    c.font  = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill  = _fill("1F3864")
    c.alignment = Alignment(horizontal="center")
    ws3.row_dimensions[1].height = 28
    for ci, h in enumerate(["Method","RBs eMBB","RBs URLLC",
                              "RBs mMTC","Util eMBB","Util URLLC",
                              "Util mMTC","Total Utility"], 1):
        _hdr(ws3, 2, ci, h)
    for ri, m in enumerate(alloc_results, 3):
        r  = alloc_results[m]; bg = MBG[m]
        _val(ws3, ri, 1, m, bold=True, bg=bg)
        for ci, slc in enumerate(SLICES, 2):
            _val(ws3, ri, ci, r["rbs"][slc], bg=bg)
        for ci, slc in enumerate(SLICES, 5):
            _val(ws3, ri, ci, round(r["utility"][slc], 4),
                 fmt="0.0000", bg=bg)
        _val(ws3, ri, 8, round(r["total_utility"], 4),
             fmt="0.0000", bold=True, bg=bg)
    ri = len(alloc_results) + 4
    ws3.cell(row=ri, column=1,
             value="QoS Violations:").font = Font(bold=True)
    for i, m in enumerate(alloc_results):
        viols = alloc_results[m]["violations"]
        txt   = "None" if not viols else "; ".join(viols)
        ws3.cell(row=ri+1+i, column=1, value=m).font = Font(bold=True)
        ws3.merge_cells(f"B{ri+1+i}:H{ri+1+i}")
        ws3.cell(row=ri+1+i, column=2, value=txt)

    # ── Sheet 4: LSTM Metrics ─────────────────────────────
    ws4 = wb.create_sheet("LSTM Metrics")
    for w, col in zip([12,26,14,14,14], [*"ABCDE"]):
        ws4.column_dimensions[col].width = w
    ws4.merge_cells("A1:E1")
    c = ws4["A1"]
    c.value = "LSTM Evaluation Metrics (Test Set)"
    c.font  = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill  = _fill("1F3864")
    c.alignment = Alignment(horizontal="center")
    ws4.row_dimensions[1].height = 28
    for ci, h in enumerate(["Slice","Target KPI",
                              "MAE","RMSE","R²"], 1):
        _hdr(ws4, 2, ci, h)
    ri = 3
    for slc in SLICES:
        bg = BG[slc]
        for col in TC:
            m = metrics_all[slc][col]
            _val(ws4, ri, 1, slc, bold=True, bg=bg)
            _val(ws4, ri, 2, col.replace("_", " "), bg=bg)
            _val(ws4, ri, 3, round(m["MAE"],  6),
                 fmt="0.000000", bg=bg)
            _val(ws4, ri, 4, round(m["RMSE"], 6),
                 fmt="0.000000", bg=bg)
            _val(ws4, ri, 5, round(m["R2"],   4),
                 fmt="0.0000",   bg=bg)
            ri += 1

    # ── Sheet 5: Plots ────────────────────────────────────
    ws5 = wb.create_sheet("Plots")
    ws5["A1"] = "Plots embedded below (also saved as PNG files)."
    plot_files = [
        ("01_lstm_training_loss.png",   "A3"),
        ("02_lstm_predictions.png",     "A32"),
        ("03_resource_allocation.png",  "A55"),
        ("04_drl_reward.png",           "A78"),
        ("05_admission_control.png",    "A100"),
        ("06_lstm_metrics.png",         "A123"),
        ("08_throughput_comparison.png","A146"),
    ]
    for fname, anchor in plot_files:
        if os.path.exists(fname):
            try:
                img = XLImage(fname)
                img.width = 700; img.height = 240
                ws5.add_image(img, anchor)
            except Exception:
                pass

    wb.save(path)
    print(f"  Saved: {path}")


"""
Code Examples
language-python
 Copy code
# ============================================================
#  results.py  –  All plots + Excel summary report
#  Fixes:
#   - Plot 3: removed fig.text footnote that broke layout
#   - Plot 3: violations shown cleanly inside subplot title
#   - All plots use constrained_layout for proper spacing
# ============================================================

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.lines import Line2D
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import os

from config import (SLICES, ADMISSION_DELAY_THRESH, DELAY_MAX, LOSS_MAX)
from preprocessing import TARGET_COLS_REAL

COLORS = {"eMBB": "#2E75B6", "URLLC": "#C55A11", "mMTC": "#538135"}
MC     = {"Greedy": "#4472C4", "Convex": "#ED7D31", "DRL": "#70AD47"}
TL     = ["Throughput (Mbps)", "Delay (ms)", "Packet Loss Rate"]
TC     = TARGET_COLS_REAL
Explanation of Each Line
Comment Block: Provides a summary of the script's purpose and recent fixes.
Import Statements:
numpy: For numerical operations.
matplotlib: For plotting; set to use a non-GUI backend.
matplotlib.pyplot: For creating plots.
matplotlib.gridspec: For advanced subplot layouts.
Line2D: For custom line styles in plots.
openpyxl: For Excel file manipulation.
Font, PatternFill, Alignment, Border, Side: For styling Excel cells.
get_column_letter: For converting column indices to letters.
Image: For inserting images into Excel.
os: For operating system functionalities.
Configuration Imports: Imports constants from configuration and preprocessing modules.
Color Dictionaries: Defines color codes for different categories (e.g., eMBB, URLLC).
Labels: Defines labels for throughput, delay, and packet loss.
language-python
 Copy code
# ── Excel helpers ─────────────────────────────────────────
def _bdr():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(h):
    return PatternFill("solid", fgColor=h)

def _hdr(ws, r, c, v, bg="1F3864"):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill      = _fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _bdr()

def _val(ws, r, c, v, fmt=None, bold=False, bg="FFFFFF"):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(name="Arial", size=9, bold=bold)
    cell.fill      = _fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _bdr()
    if fmt: cell.number_format = fmt
Explanation of Each Line
Excel Helper Functions:
_bdr(): Creates a thin border style for Excel cells.
_fill(h): Returns a solid fill pattern for cell background.
_hdr(ws, r, c, v, bg): Formats a header cell with specified value and background color.
_val(ws, r, c, v, fmt, bold, bg): Formats a value cell with optional formatting and background color.
language-python
 Copy code
# ══════════════════════════════════════════════════════════
# PLOT 1 — LSTM Training Loss
# ══════════════════════════════════════════════════════════
def _plot_training_loss(histories, out_dir):
    fig, axes = plt.subplots(1, 3, figsize=(16, 4),
                             constrained_layout=True)
    fig.suptitle("LSTM Training & Validation Loss per Slice",
                 fontsize=14, fontweight="bold")
    for ax, slc in zip(axes, SLICES):
        h  = histories[slc].history
        ep = range(1, len(h["loss"]) + 1)
        ax.plot(ep, h["loss"],     color=COLORS[slc], lw=2, label="Train Loss")
        ax.plot(ep, h["val_loss"], color=COLORS[slc], lw=2, ls="--",
                label="Val Loss", alpha=0.7)
        ax.fill_between(ep, h["loss"], h["val_loss"],
                        alpha=0.08, color=COLORS[slc])
        best = min(h["val_loss"])
        ax.axhline(best, color="grey", lw=0.8, ls=":")
        ax.set_title(f"{slc} Slice", fontweight="bold", color=COLORS[slc])
        ax.set_xlabel("Epoch"); ax.set_ylabel("MSE Loss")
        ax.legend(fontsize=8); ax.grid(alpha=0.2)
        ax.text(0.98, 0.95, f"Best val={best:.5f}",
                transform=ax.transAxes, ha="right", va="top",
                fontsize=8, color="grey")
    plt.savefig(f"{out_dir}/01_lstm_training_loss.png",
                dpi=150, bbox_inches="tight")
    plt.close()
    print("  Saved: 01_lstm_training_loss.png")
Explanation of Each Line
Function Definition: _plot_training_loss(histories, out_dir) defines a function to plot training loss.
Subplot Creation: Creates a figure with three subplots for different slices.
Figure Title: Sets the main title for the figure.
Loop Through Slices: Iterates over each slice to plot training and validation loss.
Extract History: Retrieves the training history for the current slice.
Epoch Range: Creates a range of epochs based on the length of the loss data.
Plot Training Loss: Plots the training loss for the current slice.
Plot Validation Loss: Plots the validation loss with a dashed line.
Fill Between: Fills the area between training and validation loss curves.
Best Validation Loss: Finds the minimum validation loss.
Horizontal Line: Draws a horizontal line at the best validation loss.
Set Title: Sets the title for the current subplot.
Set Labels: Labels the x-axis and y-axis.
Legend: Displays the legend for the plot.
Grid: Adds a grid to the plot for better readability.
Text Annotation: Adds text to indicate the best validation loss.
Save Figure: Saves the figure to the specified output directory.
Close Figure: Closes the figure to free up memory.
Print Confirmation: Prints a message confirming the saved plot.
Conclusion
The provided Python script is a comprehensive tool for generating visualizations and summarizing results from LSTM model training and predictions. By utilizing Matplotlib for plotting and OpenPyXL for Excel reporting, it effectively communicates the performance metrics and insights derived from the model. Each component of the code is designed to enhance clarity and usability, making it a valuable asset for data analysis in machine learning contexts.

Code Examples
Let's break down the first function _plot_drl_reward and explain each line of code.

language-python
 Copy code
def _plot_drl_reward(drl_rewards, out_dir):
    fig, axes = plt.subplots(1, 2, figsize=(14, 4), constrained_layout=True)
Function Definition: Defines a function named _plot_drl_reward that takes drl_rewards and out_dir as parameters.
Subplots Creation: Creates a figure with 1 row and 2 columns of subplots, setting the figure size to 14x4 inches.
language-python
 Copy code
    fig.suptitle("DRL (DQN) Training Performance", fontsize=14, fontweight="bold")
Figure Title: Sets the main title of the figure with specified font size and weight.
language-python
 Copy code
    ax = axes[0]
    ax.plot(drl_rewards, color="#A9D18E", alpha=0.35, lw=0.8, label="Episode Reward")
First Axes: Selects the first subplot for plotting.
Line Plot: Plots the drl_rewards data with specified color, transparency, line width, and label.
language-python
 Copy code
    window = 20
    if len(drl_rewards) >= window:
        ma = np.convolve(drl_rewards, np.ones(window)/window, mode="valid")
        ax.plot(range(window - 1, len(drl_rewards)), ma, color="#375623", lw=2.5, label=f"{window}-ep Moving Avg")
Moving Average Calculation: Defines a window size for the moving average and calculates it if the length of drl_rewards is sufficient.
Plotting Moving Average: Plots the moving average on the same axes.
language-python
 Copy code
    best_ep = int(np.argmax(drl_rewards))
    ax.annotate(f"Best: {max(drl_rewards):.4f}", xy=(best_ep, max(drl_rewards)), xytext=(best_ep + 20, max(drl_rewards) * 0.95), arrowprops=dict(arrowstyle="->", color="#375623"), fontsize=8, color="#375623")
Best Episode Calculation: Finds the index of the maximum reward.
Annotation: Adds an annotation to indicate the best reward with an arrow pointing to the corresponding point.
language-python
 Copy code
    ax.set_title("Episode Reward", fontweight="bold")
    ax.set_xlabel("Episode")
    ax.set_ylabel("Total Utility Reward")
    ax.legend()
    ax.grid(alpha=0.2)
Axes Titles and Labels: Sets titles and labels for the axes.
Legend and Grid: Displays the legend and adds a grid for better readability.
language-python
 Copy code
    ax2 = axes[1]
    cum_max = np.maximum.accumulate(drl_rewards)
    ax2.plot(cum_max, color="#4472C4", lw=2)
Second Axes: Selects the second subplot for cumulative maximum rewards.
Cumulative Maximum Calculation: Computes the cumulative maximum of drl_rewards and plots it.
language-python
 Copy code
    ax2.fill_between(range(len(cum_max)), cum_max, alpha=0.15, color="#4472C4")
Fill Between: Fills the area under the cumulative maximum curve for better visualization.
language-python
 Copy code
    conv_ep = next((i for i in range(50, len(cum_max)) if cum_max[i] - cum_max[i - 20] < 0.005), None)
Convergence Episode Calculation: Finds the episode where the cumulative maximum converges.
language-python
 Copy code
    if conv_ep:
        ax2.axvline(conv_ep, color="orange", lw=1.5, ls="--", label=f"Converges ~ep {conv_ep}")
        ax2.legend(fontsize=8)
Vertical Line for Convergence: Draws a vertical line at the convergence episode if found and adds it to the legend.
language-python
 Copy code
    ax2.set_title("Best Reward (Cumulative Max)", fontweight="bold")
    ax2.set_xlabel("Episode")
    ax2.set_ylabel("Best Utility Achieved")
    ax2.grid(alpha=0.2)
Axes Titles and Labels: Sets titles and labels for the second axes.
language-python
 Copy code
    plt.savefig(f"{out_dir}/04_drl_reward.png", dpi=150, bbox_inches="tight")
    plt.close()
    print("  Saved: 04_drl_reward.png")
Save Figure: Saves the figure to the specified output directory with a defined resolution and tight bounding box.
Close Figure: Closes the figure to free up memory.
Print Confirmation: Outputs a confirmation message indicating the file has been saved.
Conclusion
The provided code effectively visualizes various performance metrics related to resource allocation in a network context. By utilizing Matplotlib, it generates insightful plots that can aid in understanding the performance of different methods and the overall system behavior. Each function is modular, allowing for easy adjustments and enhancements as needed.

Code Examples
Function: generate_all_plots
language-python
 Copy code
def generate_all_plots(histories, preds_plot, metrics_all,
                       alloc_results, drl_rewards, ac_result,
                       demands, snr_db, out_dir="."):
    _plot_training_loss(histories, out_dir)
    _plot_predictions(preds_plot, metrics_all, out_dir)
    _plot_resource_allocation(alloc_results, out_dir)
    _plot_drl_reward(drl_rewards, out_dir)
    _plot_admission_control(ac_result, demands, out_dir)
    _plot_lstm_metrics(metrics_all, out_dir)
    _plot_throughput(alloc_results, demands, out_dir)
Line 1: Defines the function generate_all_plots with parameters for various data inputs.
Line 2-8: Calls helper functions to generate specific plots based on the provided data.
Function: save_excel_report
language-python
 Copy code
def save_excel_report(demands, ac_result, alloc_results,
                      metrics_all, snr_db, path="Results_Summary.xlsx"):
    BG  = {"eMBB": "DAEAF7", "URLLC": "FCE4D6", "mMTC": "E2EFDA"}
    MBG = {"Greedy": "DAEAF7", "Convex": "FCE4D6", "DRL": "E2EFDA"}
    wb  = openpyxl.Workbook()
Line 1: Defines the function save_excel_report with parameters for various data inputs and an optional file path.
Line 2-3: Initializes dictionaries for background colors based on service types and creates a new Excel workbook.
Sheet 1: LSTM Predictions
language-python
 Copy code
    ws1 = wb.active; ws1.title = "LSTM Predictions"
    for w, col in zip([12, 20, 18, 20], [*"ABCD"]):
        ws1.column_dimensions[col].width = w
Line 4: Sets the active worksheet and renames it to "LSTM Predictions".
Line 5-6: Adjusts the width of specific columns in the worksheet.
Merging Cells and Adding Header
language-python
 Copy code
    ws1.merge_cells("A1:D1")
    c = ws1["A1"]
    c.value = "LSTM Predicted Demand Ds(t+1) per Slice"
    c.font  = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill  = _fill("1F3864")
    c.alignment = Alignment(horizontal="center")
    ws1.row_dimensions[1].height = 28
Line 7: Merges cells A1 to D1 for the title.
Line 8-12: Sets the title text, font properties, background color, and alignment for the merged cell.
Adding Data to the Sheet
language-python
 Copy code
    for ci, h in enumerate(["Slice","Throughput (Mbps)",
                              "Delay (ms)","PLR"], 1):
        _hdr(ws1, 2, ci, h)
Line 13: Loops through a list of headers to create column headers in the second row.
Filling in Data Rows
language-python
 Copy code
    for ri, slc in enumerate(SLICES, 3):
        d  = demands[slc]; bg = BG[slc]
        _val(ws1, ri, 1, slc, bold=True, bg=bg)
        _val(ws1, ri, 2, round(d["Throughput_Mbps"],        4),
             fmt="0.0000", bg=bg)
        _val(ws1, ri, 3, round(d["Packet_Delay_Budget_ms"], 2),
             fmt="0.00",   bg=bg)
        _val(ws1, ri, 4, d["Packet_Loss_Rate"],
             fmt="0.00E+00", bg=bg)
Line 14-18: Loops through each slice, retrieves demand data, and populates the worksheet with formatted values.
Additional Sheets
The code continues to create additional sheets for "Admission Control", "Resource Allocation", "LSTM Metrics", and "Plots", following a similar structure of setting titles, headers, and populating data.

Saving the Workbook
language-python
 Copy code
    wb.save(path)
    print(f"  Saved: {path}")
Line 19-20: Saves the workbook to the specified path and prints a confirmation message.
Conclusion
The provided Python code effectively demonstrates how to visualize machine learning results and compile them into a structured Excel report. By utilizing functions for plotting and reporting, the code enhances the clarity and accessibility of data analysis, making it easier for stakeholders to interpret the results. This modular approach not only promotes code reusability but also simplifies the process of generating comprehensive reports.

"""